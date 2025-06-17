import networkx as nx
import scipy.io as scio
import math
import sys
import openpyxl
import numpy as np
import os
from utility import *
import matplotlib.pyplot as plt

orbits = 24
sat_per_orbits = 66
sats = orbits * sat_per_orbits
cycles = 5732

gs = 95

la = 25.6864
lo = -100.3357
user = lla2cbf((la, lo, 0))
print('user height',
      math.sqrt(user[0] ** 2 + user[1] ** 2 + user[2] ** 2) - 6371000.0)

delay_directory_path = '../configure/' + str(la).replace('.', '_') + '&' + str(lo).replace('.', '_')
bw_directory_path = '../configure/' + str(la).replace('.', '_') + '&' + str(lo).replace('.', '_')
os.makedirs(delay_directory_path, exist_ok=True)
os.makedirs(bw_directory_path, exist_ok=True)

# Starlink的俯角和最大仰角
depression = 44.85
elevation = 40

L = getCoverageLimitL(elevation, depression, 550 * 10 ** 3)

data = scio.loadmat('position.mat')
coordinate = data['position_cbf'][: sats]
coordinates = np.vstack(coordinate[:, 0])

ground_station_file_name = '../ground_stations.xlsx'
ground_station_llas = []
ground_station_cbfs = []
wb = openpyxl.load_workbook(ground_station_file_name)
sheet = wb.active
for i in range(2, sheet.max_row + 1):
    ground_station_llas.append((float(sheet.cell(row=i, column=1).value), float(sheet.cell(row=i, column=2).value), 0))
for ground_station_lla in ground_station_llas:
    ground_station_cbfs.append(lla2cbf(ground_station_lla))

node_nums = []
no_path_cycle_num = 0

counter = 0

selected_path = []

user_gateway_pervious = [math.inf, -1]
satellites_in_view = []

up_btlbw = []
down_btlbw = []
up_bw_record = []
down_bw_record = []
delay_record = []

for cycle in range(0, cycles):
    print("current cycle:", cycle)
    try:
        if counter % 15 == 0:
            min_path_length = math.inf
            min_path = []
            G = nx.Graph()
            edges = []
            G.add_nodes_from(range(sats + gs + 1))

            satellites_in_view = []
            achieved_gs = []

            for satellite_id in range(0, sats):
                sat_cbf = (coordinates[:, cycle][satellite_id * 3 + 0],
                           coordinates[:, cycle][satellite_id * 3 + 1],
                           coordinates[:, cycle][satellite_id * 3 + 2])
                # ISL link
                for nei_sat_id in range(0, sats):
                    if judge_neighbor_sat_node(satellite_id, nei_sat_id, orbits, sats) is False:
                        continue
                    nei_sat_cbf = (coordinates[:, cycle][nei_sat_id * 3 + 0],
                                   coordinates[:, cycle][nei_sat_id * 3 + 1],
                                   coordinates[:, cycle][nei_sat_id * 3 + 2])
                    isl_latency = computeLatency(sat_cbf, nei_sat_cbf)
                    edges.append((nei_sat_id, satellite_id, isl_latency))
                    # print('sat, nei_sat:', [satellite_id, nei_sat_id])

                # GS_sat link
                for ground_station_id in range(0, gs):
                    ground_station_cbf = ground_station_cbfs[ground_station_id]
                    gs_covered = checkSatCoverGroundStation(sat_cbf, ground_station_cbf, L)
                    if gs_covered:
                        sat_gs_latency = computeLatency(sat_cbf, ground_station_cbf)
                        achieved_gs.append([sat_gs_latency, ground_station_id + sats])
                        # remember actual ground_station_id in G has a offset of satellite_num
                        edges.append((satellite_id, ground_station_id + sats, sat_gs_latency))

                # user_sat link
                user_covered = checkSatCoverGroundStation(sat_cbf, user, L)
                if user_covered:
                    sat_latency = computeLatency(sat_cbf, user)
                    satellites_in_view.append([sat_latency, satellite_id])
                    edges.append((sats + gs, satellite_id, sat_latency))

            G.add_weighted_edges_from(edges)

            for chosen_gs in range(0, gs):
                l = nx.dijkstra_path_length(G, sats + gs, achieved_gs[chosen_gs][1])
                cur_path = nx.dijkstra_path(G, sats + gs, achieved_gs[chosen_gs][1])
                if l < min_path_length:
                    min_path_length = l
                    min_path = cur_path

            selected_path = min_path
            print('#################################15s time out')

        selected_gs = selected_path[-1] - sats
        cur_gs = ground_station_cbfs[selected_gs]

        user_cur_sat = (coordinates[:, cycle][selected_path[1] * 3 + 0],
                        coordinates[:, cycle][selected_path[1] * 3 + 1],
                        coordinates[:, cycle][selected_path[1] * 3 + 2])

        gs_cur_sat = (coordinates[:, cycle][selected_path[-2] * 3 + 0],
                      coordinates[:, cycle][selected_path[-2] * 3 + 1],
                      coordinates[:, cycle][selected_path[-2] * 3 + 2])

        if (checkSatCoverGroundStation(user_cur_sat, user, L) and checkSatCoverGroundStation(cur_gs,
                                                                                             gs_cur_sat,
                                                                                             L)) is False:
            counter = -1
            print('#################################node handover')

        user_up = calculate_bw(user_cur_sat, user, 0) / (10 ** 6)
        user_down = calculate_bw(user_cur_sat, user, 1) / (10 ** 6)
        gs_up = calculate_bw(gs_cur_sat, cur_gs, 2) / (10 ** 6)
        gs_down = calculate_bw(gs_cur_sat, cur_gs, 1) / (10 ** 6)

        sat_latency = computeLatency(user_cur_sat, user)
        sat_gs_latency = computeLatency(gs_cur_sat, cur_gs)

        if len(selected_path) == 3:
            latency = [sat_latency, sat_gs_latency]
            delay_record.append([cycle, sat_latency, sat_gs_latency])

            up_btlbw.append(min([user_up, gs_down]))
            down_btlbw.append(min([gs_up, user_down]))

            up_bw_record.append([cycle, user_up, gs_down])  # 用户上传 GS下载; GS上传 用户下载
            down_bw_record.append([cycle, gs_up, user_down])
        else:
            isl_latency = []
            isl_bw = []

            for node in range(1, len(selected_path) - 2):
                cur_sat_pre = (coordinates[:, cycle][node * 3 + 0],
                               coordinates[:, cycle][node * 3 + 1],
                               coordinates[:, cycle][node * 3 + 2])
                cur_sat_later = (coordinates[:, cycle][(node + 1) * 3 + 0],
                                 coordinates[:, cycle][(node + 1) * 3 + 1],
                                 coordinates[:, cycle][(node + 1) * 3 + 2])
                isl_latency.append(computeLatency(cur_sat_pre, cur_sat_later))
                isl_bw.append(calculate_bw(cur_sat_pre, cur_sat_later, type=1) / (10 ** 6))

            latency = [sat_latency] + isl_latency + [sat_gs_latency]
            delay_record.append(isl_latency)

            # 注意链路配置的顺序不要反了
            up_bw_record.append([user_up] + isl_bw + [gs_down])
            down_bw_record.append([user_down] + isl_bw + [gs_up])

            up_btlbw.append(min([user_up] + isl_bw + [gs_down]))
            down_btlbw.append(min([user_down] + isl_bw + [gs_up]))

        print('path:', selected_path)
        print('latency:', latency)
        print(sum(latency))

        counter += 1

    except nx.exception.NetworkXNoPath:
        print('####################### No Path #######################')

print('uplink bw:', up_btlbw)
print('down bw:', down_btlbw)

with open(bw_directory_path + '/up_bw.txt', 'w') as file:
    for row in up_bw_record:
        line = ' '.join(map(str, row))
        file.write(line + '\n')

with open(bw_directory_path + '/down_bw.txt', 'w') as file:
    for row in down_bw_record:
        line = ' '.join(map(str, row))
        file.write(line + '\n')

with open(delay_directory_path + '/latency.txt', 'w') as file:
    for row in delay_record:
        line = ' '.join(map(str, row))
        file.write(line + '\n')

plt.figure(figsize=(12, 8))
plt.plot(up_btlbw, label='uplink capacity')
plt.plot(down_btlbw, label='downlink capacity')
plt.title('ISL link bottleneck capacity', fontsize=20)  # 设置标题
plt.xlabel('cycles', fontsize=20)  # 设置X轴标签
plt.ylabel('Link Capacity(Mbps)', fontsize=20)  # 设置Y轴标签
plt.ylim(0, 210)
plt.xlim(0, 5732)
# plt.legend(loc = 'center right', fontsize=20)
plt.xticks(fontsize=15)
plt.yticks(fontsize=15)

# 显示网格线
plt.grid(True)

# 显示图形
plt.show()

x_delay = [row[0] for row in delay_record]
y_delay = [(row[1] + row[2]) * 1000 for row in delay_record]
y_delay1 = [row[1] * 1000 for row in delay_record]
y_delay2 = [row[2] * 1000 for row in delay_record]
plt.figure(figsize=(12, 8))
plt.plot(x_delay, y_delay, label='total latency')
plt.plot(x_delay, y_delay1, label='user to satellite latency')
plt.plot(x_delay, y_delay2, label='satellite to gs latency')
plt.title('Bent pipe link latency', fontsize=20)  # 设置标题
plt.xlabel('cycles', fontsize=20)  # 设置X轴标签
plt.ylabel('latency(ms)', fontsize=20)  # 设置Y轴标签
plt.xlim(0, 5732)
plt.ylim(0, 5)
# plt.legend(loc = 'lower right', fontsize=20)
plt.xticks(fontsize=15)
plt.yticks(fontsize=15)

# 显示网格线
plt.grid(True)

# 显示图形
plt.show()
