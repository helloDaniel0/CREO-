import scipy.io as scio
import numpy as np
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.animation import FuncAnimation, PillowWriter
import openpyxl
from utility import *


# 这里还有四个不要的数据
data = scio.loadmat('position.mat')
coordinate = data['position_cbf']
coordinates = np.vstack(coordinate[:-4, 0])

gs_begin = len(coordinates)

data_gs = scio.loadmat('position_with_gs.mat')
coordinates_gs = data_gs['cbfs'][gs_begin:, :]

ground_station_file_name = '../ground_stations.xlsx'
ground_station_llas = []
ground_station_cbfs = []
wb = openpyxl.load_workbook(ground_station_file_name)
sheet = wb.active
for i in range(2, sheet.max_row + 1):
    ground_station_llas.append((float(sheet.cell(row=i, column=1).value), float(sheet.cell(row=i, column=2).value), 0))
for ground_station_lla in ground_station_llas:
    ground_station_cbfs.append(lla2cbf(ground_station_lla))


orbits = 24
sat_per_orbits = 66
sats = orbits * sat_per_orbits

# 创建一个球体
phi, theta = np.mgrid[0:np.pi:50j, 0:2 * np.pi:50j]
radius = 6371000  # 新设置的半径
xs = radius * np.sin(phi) * np.cos(theta)
ys = radius * np.sin(phi) * np.sin(theta)
zs = radius * np.cos(phi)

# 创建一个三维图形
fig = plt.figure(figsize=(12, 12))

ax = fig.add_subplot(111, projection='3d')


# 绘制球体表面
ax.plot_surface(xs, ys, zs, color='c', alpha=0.15, edgecolor='none')

# 初始化固定点散点图对象
la = 42.9634
lo = -75.5244
routing_directory_path = '../configure/' + str(la).replace('.', '_') + '&' + str(lo).replace('.', '_')

routing_path = []
with open(routing_directory_path + '/routing.txt', 'r') as file:
    for lines in file:
        # 按空格分割每行数据，并转换为整数
        line_data = list(map(int, lines.split()))
        # 将数据添加到列表中
        routing_path.append(line_data)

User_lla = (la, lo, 0)
User = lla2cbf(User_lla)
x_u = User[0]
y_u = User[1]
z_u = User[2]
fixed_scatter = ax.scatter(x_u, y_u, z_u, c='k', marker='*', s=200, zorder=5)

# 创建空的散点图对象
scatter = ax.scatter([], [], [], c='b', marker='o', zorder=3)
scatter_gs = ax.scatter([], [], [], c='green', marker='^', s=50, zorder=4)
line, = ax.plot([], [], [], 'r-')

# 设置图形属性
ax.set_xlabel('X')
ax.set_ylabel('Y')
ax.set_zlabel('Z')
ax.set_title('Starlink 3D')
ax.set_box_aspect([1, 1, 1])  # 使得轴的比例相等


# 动画更新函数
def update(frame):
    # 提取当前帧的 x, y, z 数据
    x = coordinates[0::3, frame]
    y = coordinates[1::3, frame]
    z = coordinates[2::3, frame]

    x_gs = coordinates_gs[0::3, frame]
    y_gs = coordinates_gs[1::3, frame]
    z_gs = coordinates_gs[2::3, frame]

    x_line = [x_u, coordinates[routing_path[frame][1] * 3 + 0, frame],
              coordinates_gs[(routing_path[frame][2] - sats) * 3 + 0, frame]]
    y_line = [y_u, coordinates[routing_path[frame][1] * 3 + 1, frame],
              coordinates_gs[(routing_path[frame][2] - sats) * 3 + 1, frame]]
    z_line = [z_u, coordinates[routing_path[frame][1] * 3 + 2, frame],
              coordinates_gs[(routing_path[frame][2] - sats) * 3 + 2, frame]]

    line.set_data(x_line, y_line)
    line.set_3d_properties(z_line)

    # 更新散点图的数据
    scatter._offsets3d = (x, y, z)
    scatter_gs._offsets3d = (x_gs, y_gs, z_gs)
    return scatter, scatter_gs, line


# 创建动画
limited_frames = range(500)

# 创建动画
ani = FuncAnimation(fig, update, frames=limited_frames, interval=100, blit=True)

# 保存gif
# ani.save('starlink_with_gs.gif', writer=PillowWriter(fps=10))

# 显示动画
plt.show()
