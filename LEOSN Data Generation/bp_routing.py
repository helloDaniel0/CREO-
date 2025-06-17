import networkx as nx
import scipy.io as scio
import math
import sys
import openpyxl
import numpy as np
import os
from utility import *
import matplotlib.pyplot as plt

orbits = 24
sat_per_orbits = 66
sats = orbits * sat_per_orbits
cycles = 5732

gs = 95

la = 42.9634
lo = -75.5244
user = lla2cbf((la, lo, 0))

delay_directory_path = '../configure/' + str(la).replace('.', '_') + '&' + str(lo).replace('.', '_')
bw_directory_path = '../configure/' + str(la).replace('.', '_') + '&' + str(lo).replace('.', '_')
routing_directory_path = '../configure/' + str(la).replace('.', '_') + '&' + str(lo).replace('.', '_')
os.makedirs(delay_directory_path, exist_ok=True)
os.makedirs(bw_directory_path, exist_ok=True)

depression = 44.85
elevation = 40

L = getCoverageLimitL(elevation, depression, 550 * 10 ** 3)

data = scio.loadmat('position.mat')
coordinate = data['position_cbf'][: sats]
coordinates = np.vstack(coordinate[:, 0])

ground_station_file_name = '../ground_stations.xlsx'
ground_station_llas = []
ground_station_cbfs = []
wb = openpyxl.load_workbook(ground_station_file_name)
sheet = wb.active
for i in range(2, sheet.max_row + 1):
    ground_station_llas.append((float(sheet.cell(row=i, column=1).value), float(sheet.cell(row=i, column=2).value), 0))
for ground_station_lla in ground_station_llas:
    ground_station_cbfs.append(lla2cbf(ground_station_lla))

node_nums = []
no_path_cycle_num = 0

counter = 0

selected_path = []

user_gateway_pervious = [math.inf, -1]
satellites_in_view = []

up_bw = []
down_bw = []
record = []
delay_record = []
routing_record = []
switch_flag = 0

for cycle in range(0, cycles):
    print("current cycle:", cycle)
    try:
        G = nx.Graph()
        edges = []
        G.add_nodes_from(range(sats + gs + 1))

        min_path_length = math.inf
        min_path = []

        satellites_in_view = []
        achieved_gs = []

        for satellite_id in range(0, sats):
            sat_cbf = (coordinates[:, cycle][satellite_id * 3 + 0],
                       coordinates[:, cycle][satellite_id * 3 + 1],
                       coordinates[:, cycle][satellite_id * 3 + 2])

            # for ground_station_id in range(0, len(ground_station_llas)):
            #     ground_station_cbf = ground_station_cbfs[ground_station_id]
            #     gs_covered = checkSatCoverGroundStation(sat_cbf, ground_station_cbf, L)
            #     if gs_covered:
            #         sat_gs_latency = computeLatency(sat_cbf, ground_station_cbf)
            #         # remember actual ground_station_id in G has a offset of satellite_num
            #         edges.append((satellite_id, ground_station_id + sats, sat_gs_latency))
            user_covered = checkSatCoverGroundStation(sat_cbf, user, L)
            if user_covered:
                sat_latency = computeLatency(sat_cbf, user)
                satellites_in_view.append([sat_latency, satellite_id])
                edges.append((sats + gs, satellite_id, sat_latency))
        # user_gateway = user_gateway_pervious

        for sat_in_view_id in range(0, len(satellites_in_view)):
            satellite_id = satellites_in_view[sat_in_view_id][1]
            sat_cbf = (coordinates[:, cycle][satellite_id * 3 + 0],
                       coordinates[:, cycle][satellite_id * 3 + 1],
                       coordinates[:, cycle][satellite_id * 3 + 2])
            for ground_station_id in range(0, gs):
                ground_station_cbf = ground_station_cbfs[ground_station_id]
                gs_covered = checkSatCoverGroundStation(sat_cbf, ground_station_cbf, L)
                if gs_covered:
                    sat_gs_latency = computeLatency(sat_cbf, ground_station_cbf)
                    achieved_gs.append([sat_gs_latency, ground_station_id + sats])
                    # remember actual ground_station_id in G has a offset of satellite_num
                    edges.append((satellite_id, ground_station_id + sats, sat_gs_latency))

        G.add_weighted_edges_from(edges)

        for chosen_gs in range(0, len(achieved_gs)):
            l = nx.dijkstra_path_length(G, sats + gs, achieved_gs[chosen_gs][1])
            cur_path = nx.dijkstra_path(G, sats + gs, achieved_gs[chosen_gs][1])
            if l < min_path_length:
                min_path_length = l
                min_path = cur_path

        if counter % 15 == 0:
            selected_path = min_path
            print('#################################15s time out')

        selected_sat = selected_path[1]
        selected_gs = selected_path[2] - sats

        cur_sat = (coordinates[:, cycle][selected_sat * 3 + 0],
                   coordinates[:, cycle][selected_sat * 3 + 1],
                   coordinates[:, cycle][selected_sat * 3 + 2])

        cur_gs = ground_station_cbfs[selected_gs]

        if (checkSatCoverGroundStation(cur_sat, user, L) and checkSatCoverGroundStation(cur_gs, cur_sat, L)) is False:
            selected_path = min_path
            counter = 0
            selected_sat = selected_path[1]
            selected_gs = selected_path[2] - sats

            cur_sat = (coordinates[:, cycle][selected_sat * 3 + 0],
                       coordinates[:, cycle][selected_sat * 3 + 1],
                       coordinates[:, cycle][selected_sat * 3 + 2])

            cur_gs = ground_station_cbfs[selected_gs]
            print('#################################handover')

        sat_latency = computeLatency(cur_sat, user)
        sat_gs_latency = computeLatency(cur_sat, cur_gs)
        latency = [sat_latency, sat_gs_latency]
        if cycle != 0 and selected_path == routing_record[-1]:
            delay_record.append([cycle, sat_latency, sat_gs_latency, 0])
        else:
            delay_record.append([cycle, sat_latency, sat_gs_latency, 1])

        user_up = calculate_bw(cur_sat, user, 0) / (10 ** 6)
        user_down = calculate_bw(cur_sat, user, 1) / (10 ** 6)
        gs_up = calculate_bw(cur_sat, cur_gs, 2) / (10 ** 6)
        gs_down = calculate_bw(cur_sat, cur_gs, 1) / (10 ** 6)
        up_bw.append(min([user_up, gs_down]))
        down_bw.append(min([gs_up, user_down]))
        record.append([cycle, user_up, gs_down, gs_up, user_down])  # 用户上传 GS下载; GS上传 用户下载

        routing_record.append(selected_path)

        print('path:', selected_path)
        print('latency:', latency)
        print(sum(latency))

        # print(min_path)
        # print(min_path_length)

        counter += 1

    except nx.exception.NetworkXNoPath:
        print('####################### ERROR #######################')

print('uplink bw:', up_bw)
print('down bw:', down_bw)

with open(bw_directory_path + '/bw.txt', 'w') as file:
    for row in record:
        line = ' '.join(map(str, row))
        file.write(line + '\n')

with open(delay_directory_path + '/latency.txt', 'w') as file:
    for row in delay_record:
        line = ' '.join(map(str, row))
        file.write(line + '\n')

with open(delay_directory_path + '/routing.txt', 'w') as file:
    for row in routing_record:
        line = ' '.join(map(str, row))
        file.write(line + '\n')

plt.figure(figsize=(12, 8))
plt.plot(up_bw, label='uplink capacity')
plt.plot(down_bw, label='downlink capacity')
plt.title('Bent pipe link bottleneck capacity', fontsize=20)  # 设置标题
plt.xlabel('cycles', fontsize=20)  # 设置X轴标签
plt.ylabel('Link Capacity(Mbps)', fontsize=20)  # 设置Y轴标签
plt.ylim(0, 210)
plt.xlim(0, 5732)
# plt.legend(loc = 'center right', fontsize=20)
plt.xticks(fontsize=15)
plt.yticks(fontsize=15)

# 显示网格线
plt.grid(True)

# 显示图形
plt.show()

x_delay = [row[0] for row in delay_record]
y_delay = [(row[1] + row[2]) * 1000 for row in delay_record]
y_delay1 = [row[1] * 1000 for row in delay_record]
y_delay2 = [row[2] * 1000 for row in delay_record]
plt.figure(figsize=(12, 8))
plt.plot(x_delay, y_delay, label='total latency')
plt.plot(x_delay, y_delay1, label='user to satellite latency')
plt.plot(x_delay, y_delay2, label='satellite to gs latency')
plt.title('Bent pipe link latency', fontsize=20)  # 设置标题
plt.xlabel('cycles', fontsize=20)  # 设置X轴标签
plt.ylabel('latency(ms)', fontsize=20)  # 设置Y轴标签
plt.xlim(0, 5732)
plt.ylim(0, 5)
# plt.legend(loc = 'lower right', fontsize=20)
plt.xticks(fontsize=15)
plt.yticks(fontsize=15)

# 显示网格线
plt.grid(True)

# 显示图形
plt.show()
